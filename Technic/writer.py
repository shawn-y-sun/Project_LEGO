# TECHNIC/writer.py
from openpyxl.utils.cell import (
    coordinate_from_string,
    column_index_from_string,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook, Workbook
import pandas as pd
from typing import Any, List, Dict, Union, Optional

class Val:
    """
    Mini container for write specifications: holds data, target start cell, and write options.

    Attributes:
      data: Any
      start_cell: str
      orientation: 'vertical' or 'horizontal'
      include_index: bool
      include_header: bool
    """
    def __init__(
        self,
        data: Any,
        start_cell: str,
        orientation: str = 'vertical',
        include_index: bool = False,
        include_header: bool = False
    ):
        self.data = data
        self.start_cell = start_cell
        self.orientation = orientation
        self.include_index = include_index
        self.include_header = include_header

class ValueWriter:
    """
    Utility to write various Python data structures into an openpyxl Worksheet,
    starting at a specified cell, defaulting to vertical orientation.
    Supports:
      - scalar values
      - list or tuple of scalars
      - pd.Series (as vertical list)
      - pd.DataFrame (with optional headers and index)
      - dict of scalars or dict of pandas objects (writes keys in header)
    """
    def __init__(self, worksheet: Worksheet, data: Any, start_cell: str,
                 orientation: str = 'vertical',
                 include_index: bool = False, include_header: bool = False):
        self.ws = worksheet
        self.data = data
        self.start_cell = start_cell
        self.orientation = orientation
        self.include_index = include_index
        self.include_header = include_header

    def write(self, data: Any, start_cell: str,
              orientation: str = 'vertical',
              include_index: bool = False, include_header: bool = False):
        """
        Write `data` into the worksheet starting at `start_cell`.

        Parameters
        ----------
        data : scalar, list, pd.Series, pd.DataFrame, or dict
            The data to write.
        start_cell : str
            Excel cell coordinate (e.g. 'B2') to begin writing.
        orientation : {'vertical', 'horizontal'}
            Direction to write lists/Series.
        include_index : bool
            For DataFrame/Series: whether to write the index alongside data.
        include_header : bool
            For DataFrame: whether to write column headers.
        """
        col_letter, row = coordinate_from_string(start_cell)
        col = column_index_from_string(col_letter)

        def _set(cell_row, cell_col, value):
            self.ws.cell(row=cell_row, column=cell_col, value=value)

        # Scalar
        if not isinstance(data, (list, tuple, pd.Series, pd.DataFrame, dict)):
            _set(row, col, data)
            return

        # List or tuple
        if isinstance(data, (list, tuple)):
            for idx, val in enumerate(data):
                r = row + idx if orientation == 'vertical' else row
                c = col if orientation == 'vertical' else col + idx
                _set(r, c, val)
            return

        # pandas Series
        if isinstance(data, pd.Series):
            for idx, val in zip(data.index, data.values):
                offset = list(data.index).index(idx)
                r = row + offset if orientation == 'vertical' else row
                c = col if orientation == 'vertical' else col + offset
                _set(r, c, val)
            return

        # pandas DataFrame
        if isinstance(data, pd.DataFrame):
            start_r, start_c = row, col
            # Write header
            if include_header:
                for j, hdr in enumerate(data.columns):
                    _set(start_r, start_c + j, hdr)
                start_r += 1
            # Write rows
            for i, idx in enumerate(data.index):
                if include_index:
                    _set(start_r + i, start_c, idx)
                    offset = 1
                else:
                    offset = 0
                for j, col_name in enumerate(data.columns):
                    _set(start_r + i, start_c + j + offset, data.at[idx, col_name])
            return

        # dict: decide if values are scalars or pandas objects
        if isinstance(data, dict):
            # If values are all scalars
            if all(not isinstance(v, (list, tuple, pd.Series, pd.DataFrame)) for v in data.values()):
                for i, (k, v) in enumerate(data.items()):
                    _set(row + i, col, k)
                    _set(row + i, col + 1, v)
                return
            # Otherwise, iterate
            offset = 0
            for k, v in data.items():
                self.write(v,
                           get_column_letter(col + offset) + str(row),
                           orientation=orientation,
                           include_index=include_index,
                           include_header=include_header)
                offset += 2
            return

        # Fallback: write string repr
        _set(row, col, str(data))


class SheetWriter:
    """
    Coordinates multiple ValueWriter tasks on a single worksheet.
    """
    def __init__(self, worksheet: Worksheet):
        self.ws = worksheet

    def write_all(self, tasks: List[ValueWriter]):
        """
        Execute a batch of ValueWriter tasks against this sheet.
        Each task must have `data` and `start_cell` attributes.
        """
        for task in tasks:
            task.ws = self.ws
            task.write(task.data, task.start_cell,
                       orientation=task.orientation,
                       include_index=task.include_index,
                       include_header=task.include_header)


class WorkbookWriter:
    """
    Coordinates writing tasks across multiple sheets in a workbook.
    """
    def __init__(self, workbook: Workbook):
        self.wb = workbook

    def write_all(self, sheet_tasks: Dict[str, List[ValueWriter]]):
        for sheet_name, tasks in sheet_tasks.items():
            if sheet_name not in self.wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
            ws = self.wb[sheet_name]
            sw = SheetWriter(ws)
            sw.write_all(tasks)

    def save(self, path: str):
        self.wb.save(path)


class TemplateLoader:
    """
    Load and manage multiple Excel template workbooks.
    """
    def __init__(self, template_paths: List[str]):
        self.books: Dict[str, Workbook] = {
            path: load_workbook(path)
            for path in template_paths
        }

    def get_workbook(self, path: str) -> Workbook:
        return self.books[path]

    def save_all(self, output_map: Dict[str, str]):
        for in_path, out_path in output_map.items():
            wb = self.books[in_path]
            wb.save(out_path)


class TemplateWriter:
    """
    Orchestrates exporting data into multiple template workbooks.
    """
    def __init__(self, loader: TemplateLoader):
        self.loader = loader

    def write(
        self,
        sheet_tasks_map: Dict[
            str,               # template file path
            Dict[str, List[ValueWriter]]  # sheet_name -> list of tasks
        ]
    ):
        """
        Write tasks into each template workbook.

        Parameters
        ----------
        sheet_tasks_map : dict
            Mapping from template file path to a dict of sheet tasks.
        """
        for tmpl_path, sheet_tasks in sheet_tasks_map.items():
            wb = self.loader.get_workbook(tmpl_path)
            ww = WorkbookWriter(wb)
            ww.write_all(sheet_tasks)

    def save_all(self, output_map: Dict[str, str]):
        """
        Save all modified template workbooks to specified output paths.

        Parameters
        ----------
        output_map : dict
            Mapping from template input path to output path.
        """
        self.loader.save_all(output_map)
